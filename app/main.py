import io
import os
import uuid
from datetime import datetime
from typing import Optional

from authlib.integrations.starlette_client import OAuth
from fastapi import Cookie, Depends, FastAPI, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.errors import RateLimitExceeded
from slowapi.util import get_remote_address
from starlette.middleware.sessions import SessionMiddleware

from app.auth import (
    create_token,
    get_current_user,
    hash_password,
    require_admin,
    require_doctor,
    require_permission,
    verify_password,
)

from app.agents.decision_agent import evaluate_patient
from app.agents.delta_agent import compute_delta
from app.agents.discharge_agent import generate_discharge_letter
from app.agents.ingestion_agent import extract_patient_data
from app.agents.interactions_agent import check_interactions
from app.agents.validity_agent import check_medication_validity
from app.agents.summary_agent import generate_session_summary
from app.models import (
    CreateUserRequest,
    DecisionResult,
    IngestPdfRequest,
    IngestTextRequest,
    InteractionsResult,
    PatientMaster,
    PatientRecord,
    PatientTransaction,
    SaveSummaryRequest,
    SessionSummaryRequest,
    SessionSummaryResult,
    UserInfo,
    VisitDelta,
    VitalSigns,
    VitalsUpdateRequest,
)
from app.pdf_utils import extract_text_from_pdf
from app.storage import (
    SessionLocal,
    UserRow,
    consume_reset_token,
    create_reset_token,
    create_user,
    delete_user,
    get_all_records_for_export,
    get_audit_log,
    get_clinic,
    get_master,
    get_patients_by_clinic,
    get_record,
    get_record_by_internal_id,
    get_reset_token,
    get_transactions,
    get_user_by_email,
    get_user_by_id,
    get_users_by_clinic,
    list_patients,
    log_action,
    save_record,
    save_transaction,
    search_patients_by_clinic,
    seed_demo_data,
    update_user_email,
    update_user_password,
    upsert_master,
)

limiter = Limiter(key_func=get_remote_address)
app = FastAPI(title="NeoCortex AI", redirect_slashes=False)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)
app.add_middleware(SessionMiddleware, secret_key=os.environ.get("SESSION_SECRET", "dev-session-secret-change-in-prod"))
app.mount("/static", StaticFiles(directory="app/static"), name="static")

# ── Auth0 OAuth ───────────────────────────────────────────────────────────────
_auth0_domain = os.environ.get("AUTH0_DOMAIN", "")
oauth = OAuth()
if _auth0_domain:
    oauth.register(
        name="auth0",
        client_id=os.environ.get("AUTH0_CLIENT_ID"),
        client_secret=os.environ.get("AUTH0_CLIENT_SECRET"),
        server_metadata_url=f"https://{_auth0_domain}/.well-known/openid-configuration",
        client_kwargs={"scope": "openid email profile"},
    )


@app.on_event("startup")
def on_startup():
    with SessionLocal() as session:
        seed_demo_data(session)


# ── Auth routes ──────────────────────────────────────────────────────────────

@app.get("/login")
async def login_page() -> FileResponse:
    return FileResponse("app/static/login.html")


@app.post("/auth/login")
@limiter.limit("10/minute")
async def do_login(request: Request, body: dict):
    id_number = body.get("id_number", "")
    password = body.get("password", "")
    with SessionLocal() as session:
        user = get_user_by_id(session, id_number)
        if not user or not verify_password(password, user.hashed_password):
            raise HTTPException(status_code=401, detail="מספר תעודת זהות או סיסמה שגויים")
        import json as _json
        perms = _json.loads(user.permissions) if user.permissions else []
        token_data = {
            "id_number": user.id_number,
            "full_name": user.full_name,
            "role": user.role,
            "clinic_id": user.clinic_id,
            "specialty": user.specialty,
            "permissions": perms,
        }
        token = create_token(token_data)
        log_action(user.id_number, "login", user_name=user.full_name, clinic_id=user.clinic_id)
        response = JSONResponse(content={
            "id_number": user.id_number,
            "full_name": user.full_name,
            "role": user.role,
            "clinic_id": user.clinic_id,
            "specialty": user.specialty,
            "permissions": perms,
        })
        response.set_cookie("neocortex_token", token, httponly=True, samesite="lax", max_age=28800)
        return response


@app.post("/auth/logout")
async def logout():
    response = JSONResponse(content={"ok": True})
    response.delete_cookie("neocortex_token")
    return response


@app.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return user


@app.post("/auth/change-password")
async def change_password(body: dict, user: dict = Depends(get_current_user)):
    current = body.get("current_password", "")
    new_pw = body.get("new_password", "")
    if not current or not new_pw:
        raise HTTPException(status_code=400, detail="נדרשות סיסמה נוכחית וסיסמה חדשה")
    if len(new_pw) < 6:
        raise HTTPException(status_code=400, detail="הסיסמה החדשה חייבת להכיל לפחות 6 תווים")
    with SessionLocal() as session:
        db_user = get_user_by_id(session, user["id_number"])
        if not db_user or not verify_password(current, db_user.hashed_password):
            raise HTTPException(status_code=401, detail="הסיסמה הנוכחית שגויה")
        update_user_password(session, user["id_number"], hash_password(new_pw))
    log_action(user["id_number"], "change_password", user_name=user.get("full_name"), clinic_id=user.get("clinic_id"))
    return {"ok": True}


@app.post("/auth/forgot-password")
@limiter.limit("5/minute")
async def forgot_password(request: Request, body: dict):
    id_number = body.get("id_number", "").strip()
    email = body.get("email", "").strip().lower()
    if not id_number or not email:
        raise HTTPException(status_code=400, detail="נדרשים תעודת זהות ואימייל")
    with SessionLocal() as session:
        user = get_user_by_id(session, id_number)
        # Always return OK to avoid user enumeration
        if not user or (user.email or "").lower() != email:
            return {"ok": True, "msg": "אם הפרטים נכונים, נשלח אליך מייל לאיפוס סיסמה"}
        token = create_reset_token(session, id_number)
        _send_reset_email(email, user.full_name, token)
    return {"ok": True, "msg": "אם הפרטים נכונים, נשלח אליך מייל לאיפוס סיסמה"}


@app.post("/auth/reset-password")
async def reset_password(body: dict):
    token = body.get("token", "").strip()
    new_pw = body.get("new_password", "")
    if not token or not new_pw:
        raise HTTPException(status_code=400, detail="נדרשים טוקן וסיסמה חדשה")
    if len(new_pw) < 6:
        raise HTTPException(status_code=400, detail="הסיסמה חייבת להכיל לפחות 6 תווים")
    with SessionLocal() as session:
        row = get_reset_token(session, token)
        if not row:
            raise HTTPException(status_code=400, detail="הקישור פג תוקף או כבר נוצל")
        update_user_password(session, row.id_number, hash_password(new_pw))
        consume_reset_token(session, token)
    return {"ok": True}


@app.patch("/admin/users/{id_number}/email")
async def set_user_email(id_number: str, body: dict, user: dict = Depends(require_admin)):
    email = body.get("email", "").strip().lower()
    if not email or "@" not in email:
        raise HTTPException(status_code=400, detail="כתובת אימייל לא תקינה")
    with SessionLocal() as session:
        ok = update_user_email(session, id_number, email)
        if not ok:
            raise HTTPException(status_code=404, detail="משתמש לא נמצא")
    return {"ok": True}


def _send_reset_email(to_email: str, full_name: str, token: str) -> None:
    import smtplib
    from email.mime.text import MIMEText
    smtp_host = os.environ.get("SMTP_HOST", "")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    smtp_user = os.environ.get("SMTP_USER", "")
    smtp_pass = os.environ.get("SMTP_PASS", "")
    from_email = os.environ.get("SMTP_FROM", smtp_user)
    base_url = os.environ.get("BASE_URL", "https://neocortex-api.onrender.com")
    if not smtp_host or not smtp_user:
        # Dev mode: just print the link
        print(f"[DEV] Password reset link for {to_email}: {base_url}/reset-password?token={token}")
        return
    reset_url = f"{base_url}/reset-password?token={token}"
    body = f"""שלום {full_name},

קיבלנו בקשה לאיפוס הסיסמה שלך במערכת NeoCortex AI.

לאיפוס הסיסמה לחץ על הקישור:
{reset_url}

הקישור תקף לשעה אחת.

אם לא ביקשת לאפס סיסמה, התעלם ממייל זה.

NeoCortex AI
"""
    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = "איפוס סיסמה — NeoCortex AI"
    msg["From"] = from_email
    msg["To"] = to_email
    with smtplib.SMTP(smtp_host, smtp_port) as server:
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.sendmail(from_email, [to_email], msg.as_string())


# Legacy login/logout (form-based, kept for compatibility)
@app.post("/login")
async def do_login_legacy(username: str = Form(...), password: str = Form(...)):
    with SessionLocal() as session:
        user = get_user_by_id(session, username)
        if not user or not verify_password(password, user.hashed_password):
            return RedirectResponse("/login?error=1", status_code=303)
        token_data = {
            "id_number": user.id_number,
            "full_name": user.full_name,
            "role": user.role,
            "clinic_id": user.clinic_id,
            "specialty": user.specialty,
        }
        token = create_token(token_data)
        response = RedirectResponse("/", status_code=303)
        response.set_cookie("neocortex_token", token, httponly=True, samesite="lax", max_age=28800)
        return response


@app.post("/logout")
async def logout_legacy():
    response = RedirectResponse("/login", status_code=303)
    response.delete_cookie("neocortex_token")
    return response


@app.get("/me")
async def me_legacy(user: dict = Depends(get_current_user)):
    return user


# ── Pages ────────────────────────────────────────────────────────────────────

@app.get("/")
async def index(neocortex_token: Optional[str] = Cookie(default=None)) -> FileResponse:
    if not neocortex_token:
        return RedirectResponse("/login")
    return FileResponse("app/static/index.html")


@app.get("/auth/auth0/login")
async def auth0_login(request: Request):
    if not _auth0_domain:
        raise HTTPException(status_code=503, detail="Auth0 not configured")
    callback_url = os.environ.get("AUTH0_CALLBACK_URL", "https://neocortex-api.onrender.com/auth/auth0/callback")
    return await oauth.auth0.authorize_redirect(request, callback_url)


@app.get("/auth/auth0/callback")
async def auth0_callback(request: Request):
    if not _auth0_domain:
        raise HTTPException(status_code=503, detail="Auth0 not configured")
    try:
        token = await oauth.auth0.authorize_access_token(request)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Auth0 error: {e}")

    userinfo = token.get("userinfo") or {}
    email = (userinfo.get("email") or "").lower().strip()
    if not email:
        raise HTTPException(status_code=400, detail="לא נמצא אימייל בחשבון Auth0")

    with SessionLocal() as session:
        user = get_user_by_email(session, email)
        if not user:
            return RedirectResponse("/login?error=auth0_no_user")
        import json as _json
        perms = _json.loads(user.permissions) if user.permissions else []
        token_data = {
            "id_number": user.id_number,
            "full_name": user.full_name,
            "role": user.role,
            "clinic_id": user.clinic_id,
            "specialty": user.specialty,
            "permissions": perms,
        }
        jwt_token = create_token(token_data)
        log_action(user.id_number, "login_auth0", user_name=user.full_name, clinic_id=user.clinic_id)

    redirect_to = "/" if user.role != "admin" else "/admin"
    response = RedirectResponse(redirect_to)
    response.set_cookie("neocortex_token", jwt_token, httponly=True, samesite="lax", max_age=28800)
    return response


@app.get("/auth/auth0/logout")
async def auth0_logout(request: Request):
    response = RedirectResponse("/login")
    response.delete_cookie("neocortex_token")
    if _auth0_domain:
        base_url = os.environ.get("BASE_URL", "https://neocortex-api.onrender.com")
        logout_url = (
            f"https://{_auth0_domain}/v2/logout"
            f"?client_id={os.environ.get('AUTH0_CLIENT_ID')}"
            f"&returnTo={base_url}/login"
        )
        return RedirectResponse(logout_url)
    return response


@app.get("/forgot-password")
async def forgot_password_page() -> FileResponse:
    return FileResponse("app/static/forgot-password.html")


@app.get("/reset-password")
async def reset_password_page() -> FileResponse:
    return FileResponse("app/static/reset-password.html")


@app.get("/debug-seed")
async def debug_seed():
    from app.storage import SessionLocal, seed_demo_data, UserRow
    try:
        with SessionLocal() as session:
            has_seed_user = session.get(UserRow, "000000000")
            all_users = session.query(UserRow).all()
            seed_demo_data(session)
            all_users_after = session.query(UserRow).all()
        return {
            "had_seed_user_before": bool(has_seed_user),
            "users_before": [u.id_number for u in all_users],
            "users_after": [u.id_number for u in all_users_after],
        }
    except Exception as e:
        return {"error": str(e)}


@app.get("/setup-admin")
async def setup_admin(secret: str = ""):
    if secret != os.environ.get("SETUP_SECRET", ""):
        raise HTTPException(status_code=403, detail="Forbidden")
    with SessionLocal() as session:
        from app.storage import ClinicRow
        clinic = session.query(ClinicRow).filter(ClinicRow.id == "default").first()
        if not clinic:
            session.add(ClinicRow(id="default", name="מרפאת ארבל"))
        else:
            clinic.name = "מרפאת ארבל"
        session.commit()
        # Move all clinic-demo users to default clinic
        session.query(UserRow).filter_by(clinic_id="clinic-demo").update({"clinic_id": "default"})
        # Move demo patient to default clinic
        from app.storage import PatientMasterRow, PatientRecordRow
        session.query(PatientMasterRow).filter_by(clinic_id="clinic-demo").update({"clinic_id": "default"})
        session.query(PatientRecordRow).filter_by(clinic_id="clinic-demo").update({"clinic_id": "default"})
        session.commit()

        existing = session.query(UserRow).filter(UserRow.id_number == "999735372").first()
        if not existing:
            existing = session.query(UserRow).filter(UserRow.id_number == "999999999").first()
        if existing:
            existing.id_number = "999735372"
            existing.email = "ferrerashirel@gmail.com"
            existing.role = "admin"
            existing.full_name = "עברי שמעון"
            existing.specialty = "רפואת משפחה"
            existing.clinic_id = "default"
            session.commit()
            return {"status": "updated", "user": "עברי שמעון", "id": "999735372"}
        user = UserRow(
            id_number="999735372",
            hashed_password=hash_password("NeoCortex2026!"),
            role="admin",
            full_name="עברי שמעון",
            specialty="רפואת משפחה",
            clinic_id="default",
            email="ferrerashirel@gmail.com",
        )
        session.add(user)
        session.commit()
    return {"status": "created", "user": "עברי שמעון", "id": "999999999", "password": "NeoCortex2026!"}


@app.get("/admin")
async def admin_page(user: dict = Depends(require_admin)) -> FileResponse:
    return FileResponse("app/static/admin.html")


# ── Admin API ────────────────────────────────────────────────────────────────

@app.get("/admin/clinic")
async def get_clinic_info(user: dict = Depends(require_admin)):
    with SessionLocal() as session:
        clinic = get_clinic(session, user["clinic_id"])
        if not clinic:
            raise HTTPException(status_code=404, detail="קליניקה לא נמצאה")
        return {"id": clinic.id, "name": clinic.name}


@app.get("/admin/audit-log")
async def admin_audit_log(user: dict = Depends(require_admin)):
    return get_audit_log(user["clinic_id"])


@app.get("/admin/users")
async def list_users(user: dict = Depends(require_admin)):
    import json as _json
    with SessionLocal() as session:
        users = get_users_by_clinic(session, user["clinic_id"])
        return [
            {
                "id_number": u.id_number,
                "full_name": u.full_name,
                "specialty": u.specialty,
                "role": u.role,
                "clinic_id": u.clinic_id,
                "permissions": _json.loads(u.permissions) if u.permissions else [],
            }
            for u in users
        ]


@app.post("/admin/users")
async def add_user(req: CreateUserRequest, user: dict = Depends(require_admin)):
    import json as _json
    if req.role not in ("doctor", "secretary", "admin", "nurse", "intern"):
        raise HTTPException(status_code=400, detail="תפקיד לא תקין")
    with SessionLocal() as session:
        existing = get_user_by_id(session, req.id_number)
        if existing:
            raise HTTPException(status_code=409, detail="משתמש עם תעודת זהות זו כבר קיים")
        hashed = hash_password(req.password)
        permissions = req.permissions if req.permissions is not None else None
        new_user = create_user(
            session,
            id_number=req.id_number,
            full_name=req.full_name,
            specialty=req.specialty,
            role=req.role,
            clinic_id=user["clinic_id"],
            hashed_password=hashed,
            permissions=permissions,
        )
        if req.email:
            update_user_email(session, req.id_number, req.email)
        return {
            "id_number": new_user.id_number,
            "full_name": new_user.full_name,
            "specialty": new_user.specialty,
            "role": new_user.role,
            "clinic_id": new_user.clinic_id,
            "permissions": _json.loads(new_user.permissions) if new_user.permissions else [],
        }


@app.delete("/admin/users/{id_number}")
async def remove_user(id_number: str, user: dict = Depends(require_admin)):
    # Prevent self-deletion
    if id_number == user.get("id_number"):
        raise HTTPException(status_code=400, detail="לא ניתן למחוק את המשתמש הנוכחי")
    with SessionLocal() as session:
        ok = delete_user(session, id_number)
        if not ok:
            raise HTTPException(status_code=404, detail="משתמש לא נמצא")
        return {"ok": True}


# ── Ingest helpers ────────────────────────────────────────────────────────────

def _build_transaction(record: PatientRecord, raw_text: str) -> PatientTransaction:
    today = datetime.utcnow().strftime("%Y-%m-%d")
    return PatientTransaction(
        transaction_id=str(uuid.uuid4()),
        patient_id=record.patient_id,
        date=today,
        transaction_type="referral",
        raw_text=raw_text,
        extracted=record,
    )


def _user_tags(user: dict) -> dict:
    clinic_id = user.get("clinic_id")
    doctor_id = user.get("id_number") if user.get("role") == "doctor" else None
    specialty = user.get("specialty") if user.get("role") == "doctor" else None
    return {"clinic_id": clinic_id, "doctor_id_number": doctor_id, "specialty": specialty}


# ── Ingest routes ─────────────────────────────────────────────────────────────

@app.post("/ingest/pdf", response_model=PatientTransaction)
@limiter.limit("30/minute")
async def ingest_pdf(request: Request, patient_id: str, file: UploadFile, user: dict = Depends(require_permission("edit_records"))) -> PatientTransaction:
    file_bytes = await file.read()
    raw_text = extract_text_from_pdf(file_bytes)
    record = extract_patient_data(patient_id, raw_text, source="pdf")
    tags = _user_tags(user)
    save_record(record, **tags)
    upsert_master(patient_id, record.full_name, record.date_of_birth, record.gender)
    tx = _build_transaction(record, raw_text)
    save_transaction(tx, clinic_id=tags["clinic_id"], doctor_id_number=tags["doctor_id_number"])
    return tx


@app.post("/ingest/pdf-base64", response_model=PatientTransaction)
async def ingest_pdf_base64(request: IngestPdfRequest, user: dict = Depends(require_permission("edit_records"))) -> PatientTransaction:
    import base64
    file_bytes = base64.b64decode(request.pdf_base64)
    raw_text = extract_text_from_pdf(file_bytes)
    record = extract_patient_data(request.patient_id, raw_text, source="pdf")
    tags = _user_tags(user)
    save_record(record, **tags)
    upsert_master(record.patient_id, record.full_name, record.date_of_birth, record.gender)
    tx = _build_transaction(record, raw_text)
    save_transaction(tx, clinic_id=tags["clinic_id"], doctor_id_number=tags["doctor_id_number"])
    return tx


@app.post("/ingest/text", response_model=PatientTransaction)
@limiter.limit("30/minute")
async def ingest_text(request: Request, body: IngestTextRequest, user: dict = Depends(require_permission("edit_records"))) -> PatientTransaction:
    record = extract_patient_data(body.patient_id, body.text, source="text")
    tags = _user_tags(user)
    save_record(record, **tags)
    upsert_master(record.patient_id, record.full_name, record.date_of_birth, record.gender)
    tx = _build_transaction(record, body.text)
    save_transaction(tx, clinic_id=tags["clinic_id"], doctor_id_number=tags["doctor_id_number"])
    log_action(user["id_number"], "ingest", user_name=user.get("full_name"),
               clinic_id=tags["clinic_id"], patient_id=body.patient_id)
    return tx


# ── Patient routes ────────────────────────────────────────────────────────────

@app.get("/patients/search")
async def search_patients(q: str, user: dict = Depends(require_permission("view_records"))) -> list[dict]:
    clinic_id = user.get("clinic_id")
    if not clinic_id or not q.strip():
        return []
    with SessionLocal() as session:
        results = search_patients_by_clinic(session, clinic_id, q)
    log_action(user["id_number"], "search_patients", user_name=user.get("full_name"),
               clinic_id=clinic_id, detail=q)
    return results


@app.get("/patients/export/excel")
async def export_patients_excel(user: dict = Depends(require_permission("view_records"))):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    clinic_id = user.get("clinic_id")
    if not clinic_id:
        raise HTTPException(status_code=400, detail="אין מרפאה מוגדרת")
    with SessionLocal() as session:
        records = get_all_records_for_export(session, clinic_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "מטופלים"
    ws.sheet_view.rightToLeft = True

    headers = ["שם מלא", "ת.ז", "תאריך לידה", "מגדר", "תלונה עיקרית",
               "תרופות", "אלרגיות", "היסטוריה רפואית", "תסמינים"]
    header_fill = PatternFill("solid", fgColor="1A56DB")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="right")

    for row_idx, rec in enumerate(records, 2):
        meds = ", ".join(rec.get("medications") or [])
        allergies = ", ".join(rec.get("allergies") or [])
        history = rec.get("medical_history") or []
        history_str = ", ".join(
            (c.get("name") if isinstance(c, dict) else str(c)) for c in history
        )
        symptoms = ", ".join(rec.get("symptoms") or [])
        row_data = [
            rec.get("full_name", ""),
            rec.get("patient_id", ""),
            rec.get("date_of_birth", ""),
            rec.get("gender", ""),
            rec.get("chief_complaint", ""),
            meds, allergies, history_str, symptoms,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val or "")
            cell.alignment = Alignment(horizontal="right")
        if row_idx % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = PatternFill("solid", fgColor="F0F4FF")

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    log_action(user["id_number"], "export_excel", user_name=user.get("full_name"), clinic_id=clinic_id)
    from datetime import date
    filename = f"neocortex_patients_{date.today()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/patients")
async def get_patients(user: dict = Depends(get_current_user)) -> list[dict]:
    clinic_id = user.get("clinic_id")
    if not clinic_id:
        # Fallback: return all (legacy)
        return [
            {"patient_id": p.patient_id, "full_name": p.full_name,
             "date_of_birth": p.date_of_birth, "gender": p.gender}
            for p in list_patients()
        ]
    with SessionLocal() as session:
        rows = get_patients_by_clinic(session, clinic_id)
        if user.get("role") == "doctor":
            specialty = user.get("specialty")
            rows = [r for r in rows if r.specialty is None or r.specialty == specialty]
        import json
        result = []
        for row in rows:
            data = json.loads(row.data) if isinstance(row.data, str) else row.data
            result.append({
                "patient_id": data.get("patient_id", row.patient_id),
                "full_name": data.get("full_name"),
                "date_of_birth": data.get("date_of_birth"),
                "gender": data.get("gender"),
            })
        return result


@app.get("/patients/{patient_id}", response_model=PatientRecord)
async def get_patient(patient_id: str, user: dict = Depends(require_permission("view_records"))) -> PatientRecord:
    record = get_record(patient_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    log_action(user["id_number"], "view_patient", user_name=user.get("full_name"),
               clinic_id=user.get("clinic_id"), patient_id=patient_id)
    return record


@app.get("/patients/{patient_id}/transactions", response_model=list[PatientTransaction])
async def get_patient_transactions(patient_id: str, user: dict = Depends(require_permission("view_records"))) -> list[PatientTransaction]:
    return get_transactions(patient_id)


@app.get("/patients/{patient_id}/print")
async def print_patient(patient_id: str, user: dict = Depends(require_permission("view_records"))):
    record = get_record(patient_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")

    def row(label: str, value) -> str:
        if not value:
            return ""
        import html as html_mod
        return f'<div class="field"><span class="label">{html_mod.escape(label)}</span><span class="value">{html_mod.escape(str(value))}</span></div>'

    def section(title: str, content: str) -> str:
        if not content.strip():
            return ""
        return f'<section><h2>{title}</h2>{content}</section>'

    def bullet_list(items) -> str:
        import html as html_mod
        if not items:
            return ""
        return "<ul>" + "".join(f"<li>{html_mod.escape(i)}</li>" for i in items) + "</ul>"

    import html as html_mod

    vitals_html = ""
    if record.vitals:
        v = record.vitals
        vitals_html = section("מדדים חיוניים", (
            row("דופק", f"{v.heart_rate} bpm" if v.heart_rate else None) +
            row("לחץ דם", f"{v.blood_pressure_systolic}/{v.blood_pressure_diastolic} mmHg" if v.blood_pressure_systolic and v.blood_pressure_diastolic else None) +
            row("טמפרטורה", f"{v.temperature_celsius} °C" if v.temperature_celsius else None) +
            row("חמצן בדם", f"{v.spo2_percent}%" if v.spo2_percent else None)
        ))

    now = datetime.now().strftime("%d/%m/%Y %H:%M")
    html_content = f"""<!DOCTYPE html>
<html lang="he" dir="rtl">
<head>
<meta charset="UTF-8">
<title>תיק מטופל — {html_mod.escape(patient_id)}</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Arial, Helvetica, sans-serif; font-size: 13px; color: #111; direction: rtl; padding: 24px; }}
  h1 {{ font-size: 20px; margin-bottom: 4px; }}
  .meta {{ color: #666; font-size: 11px; margin-bottom: 20px; }}
  section {{ margin-bottom: 18px; page-break-inside: avoid; }}
  h2 {{ font-size: 14px; color: #1a56db; border-bottom: 1px solid #e5e7eb; padding-bottom: 3px; margin-bottom: 8px; }}
  .field {{ display: flex; gap: 8px; margin-bottom: 5px; }}
  .label {{ color: #6b7280; min-width: 120px; font-size: 11px; }}
  .value {{ flex: 1; }}
  ul {{ padding-right: 18px; }}
  li {{ margin-bottom: 3px; }}
  .disclaimer {{ font-size: 10px; color: #9ca3af; border-top: 1px solid #e5e7eb; margin-top: 24px; padding-top: 8px; }}
  @media print {{
    body {{ padding: 0; }}
    button {{ display: none; }}
  }}
</style>
</head>
<body>
<h1>NeoCortex AI — תיק מטופל</h1>
<div class="meta">ת.ז: {html_mod.escape(patient_id)} &nbsp;|&nbsp; הופק: {now}</div>
{section("פרטים אישיים", row("שם מלא", record.full_name) + row("תאריך לידה", record.date_of_birth) + row("מין", record.gender) + row("תלונה עיקרית", record.chief_complaint))}
{section("תרופות", bullet_list(record.medications))}
{vitals_html}
{section("אלרגיות", bullet_list(record.allergies))}
{section("היסטוריה רפואית", bullet_list(record.medical_history))}
{section("תסמינים", bullet_list(record.symptoms))}
<div class="disclaimer">מסמך זה הופק למטרות עיון בלבד ואינו תחליף לשיפוט קליני מקצועי.</div>
<script>window.onload = function() {{ window.print(); }}</script>
</body>
</html>"""

    log_action(user["id_number"], "export_pdf", user_name=user.get("full_name"),
               clinic_id=user.get("clinic_id"), patient_id=patient_id)
    return Response(content=html_content, media_type="text/html; charset=utf-8")


@app.patch("/patients/{patient_id}/vitals", response_model=PatientRecord)
async def update_vitals(patient_id: str, vitals: VitalsUpdateRequest, user: dict = Depends(require_permission("edit_records"))) -> PatientRecord:
    record = get_record(patient_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    current = record.vitals or VitalSigns()
    updated = current.model_copy(update={k: v for k, v in vitals.model_dump().items() if v is not None})
    record.vitals = updated
    save_record(record)
    return record


@app.post("/patients/{patient_id}/session-summary", response_model=SessionSummaryResult)
async def session_summary(patient_id: str, request: SessionSummaryRequest, user: dict = Depends(require_permission("session_summary"))) -> SessionSummaryResult:
    record = get_record(patient_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    if not request.notes or not request.notes.strip():
        raise HTTPException(status_code=400, detail="Notes cannot be empty")
    transactions = get_transactions(patient_id)
    previous_summary = None
    if len(transactions) > 1:
        prev = transactions[1].extracted
        previous_summary = prev.chief_complaint
    summary = generate_session_summary(
        patient_name=record.full_name or patient_id,
        notes=request.notes,
        previous_summary=previous_summary,
    )
    return SessionSummaryResult(patient_id=patient_id, summary=summary)


@app.post("/patients/{patient_id}/save-summary", response_model=PatientTransaction)
async def save_summary(patient_id: str, request: SaveSummaryRequest, user: dict = Depends(require_permission("session_summary"))) -> PatientTransaction:
    record = get_record(patient_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    today = datetime.utcnow().strftime("%Y-%m-%d")
    note = f"סיכום ביקור {today}"
    if request.doctor_name:
        note += f" | ד\"ר {request.doctor_name}"
    note += f"\n\n{request.summary}"
    visit_record = record.model_copy(update={"chief_complaint": note, "source": "visit"})
    tx = PatientTransaction(
        transaction_id=str(uuid.uuid4()),
        patient_id=patient_id,
        date=today,
        transaction_type="visit",
        raw_text=request.summary,
        extracted=visit_record,
    )
    save_transaction(tx)
    return tx


@app.post("/decision/{patient_id}", response_model=DecisionResult)
async def run_decision(patient_id: str, user: dict = Depends(require_permission("clinical_analysis"))) -> DecisionResult:
    record = get_record(patient_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    transactions = get_transactions(patient_id)
    history = transactions[1:] if len(transactions) > 1 else []
    log_action(user["id_number"], "clinical_analysis", user_name=user.get("full_name"),
               clinic_id=user.get("clinic_id"), patient_id=patient_id)
    result = evaluate_patient(record, history=history)
    if len(transactions) >= 2:
        result.visit_delta = compute_delta(record, transactions[1].extracted)
    return result


@app.post("/patients/{patient_id}/interactions", response_model=InteractionsResult)
async def run_interactions(
    patient_id: str,
    user: dict = Depends(require_permission("drug_interactions")),
) -> InteractionsResult:
    record = get_record(patient_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    log_action(user["id_number"], "drug_interactions", user_name=user.get("full_name"),
               clinic_id=user.get("clinic_id"), patient_id=patient_id)
    return check_interactions(patient_id, record.medications)


# ── UUID-based routes (internal_id) — avoids exposing national ID in URLs ────

@app.get("/p/{internal_id}", response_model=PatientRecord)
async def get_patient_by_internal_id(internal_id: str, user: dict = Depends(require_permission("view_records"))) -> PatientRecord:
    record = get_record_by_internal_id(internal_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    log_action(user["id_number"], "view_patient", user_name=user.get("full_name"),
               clinic_id=user.get("clinic_id"), patient_id=record.patient_id)
    return record


@app.get("/p/{internal_id}/print")
async def print_patient_by_internal_id(internal_id: str, user: dict = Depends(require_permission("view_records"))):
    record = get_record_by_internal_id(internal_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    # Delegate to the existing print logic by calling it directly
    return await print_patient(record.patient_id, user)


@app.post("/p/{internal_id}/decision", response_model=DecisionResult)
async def run_decision_by_internal_id(internal_id: str, user: dict = Depends(require_permission("clinical_analysis"))) -> DecisionResult:
    record = get_record_by_internal_id(internal_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    transactions = get_transactions(record.patient_id)
    history = transactions[1:] if len(transactions) > 1 else []
    log_action(user["id_number"], "clinical_analysis", user_name=user.get("full_name"),
               clinic_id=user.get("clinic_id"), patient_id=record.patient_id)
    result = evaluate_patient(record, history=history)
    if len(transactions) >= 2:
        result.visit_delta = compute_delta(record, transactions[1].extracted)
    return result


@app.post("/p/{internal_id}/interactions", response_model=InteractionsResult)
async def run_interactions_by_internal_id(
    internal_id: str,
    user: dict = Depends(require_permission("drug_interactions")),
) -> InteractionsResult:
    record = get_record_by_internal_id(internal_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    log_action(user["id_number"], "drug_interactions", user_name=user.get("full_name"),
               clinic_id=user.get("clinic_id"), patient_id=record.patient_id)
    return check_interactions(record.patient_id, record.medications)


@app.post("/p/{internal_id}/medication-validity")
async def check_validity_by_internal_id(
    internal_id: str,
    user: dict = Depends(require_permission("drug_interactions")),
):
    record = get_record_by_internal_id(internal_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    referral_date = record.referral_date or (record.created_at.strftime("%Y-%m-%d") if record.created_at else None)
    return check_medication_validity(record.patient_id, record.medications, referral_date)


@app.patch("/p/{internal_id}/vitals", response_model=PatientRecord)
async def update_vitals_by_internal_id(internal_id: str, vitals: VitalsUpdateRequest, user: dict = Depends(require_permission("edit_records"))) -> PatientRecord:
    record = get_record_by_internal_id(internal_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    current = record.vitals or VitalSigns()
    updated = current.model_copy(update={k: v for k, v in vitals.model_dump().items() if v is not None})
    record.vitals = updated
    save_record(record)
    return record


@app.patch("/patients/{patient_id}/conditions", response_model=PatientRecord)
async def update_conditions(patient_id: str, body: dict, user: dict = Depends(require_permission("edit_records"))) -> PatientRecord:
    record = get_record(patient_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    from app.models import MedicalCondition
    conditions = body.get("conditions", [])
    record = record.model_copy(update={"medical_history": [
        MedicalCondition(**c) if isinstance(c, dict) else c for c in conditions
    ]})
    save_record(record)
    return record


@app.patch("/p/{internal_id}/conditions", response_model=PatientRecord)
async def update_conditions_by_internal_id(internal_id: str, body: dict, user: dict = Depends(require_permission("edit_records"))) -> PatientRecord:
    record = get_record_by_internal_id(internal_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    from app.models import MedicalCondition
    conditions = body.get("conditions", [])
    record = record.model_copy(update={"medical_history": [
        MedicalCondition(**c) if isinstance(c, dict) else c for c in conditions
    ]})
    save_record(record)
    return record


@app.post("/p/{internal_id}/session-summary", response_model=SessionSummaryResult)
async def session_summary_by_internal_id(internal_id: str, request: SessionSummaryRequest, user: dict = Depends(require_permission("session_summary"))) -> SessionSummaryResult:
    record = get_record_by_internal_id(internal_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    if not request.notes or not request.notes.strip():
        raise HTTPException(status_code=400, detail="Notes cannot be empty")
    patient_id = record.patient_id
    transactions = get_transactions(patient_id)
    previous_summary = None
    if len(transactions) > 1:
        prev = transactions[1].extracted
        previous_summary = prev.chief_complaint
    summary = generate_session_summary(
        patient_name=record.full_name or patient_id,
        notes=request.notes,
        previous_summary=previous_summary,
    )
    return SessionSummaryResult(patient_id=patient_id, summary=summary)


@app.post("/p/{internal_id}/discharge-letter")
async def discharge_letter_by_internal_id(internal_id: str, user: dict = Depends(require_permission("clinical_analysis"))):
    record = get_record_by_internal_id(internal_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    doctor_name = user.get("full_name") if user.get("role") == "doctor" else None
    doctor_specialty = user.get("specialty") if user.get("role") == "doctor" else None
    log_action(user["id_number"], "discharge_letter", user_name=user.get("full_name"),
               clinic_id=user.get("clinic_id"), patient_id=record.patient_id)
    result = generate_discharge_letter(record, doctor_name=doctor_name, doctor_specialty=doctor_specialty)
    return result


@app.post("/p/{internal_id}/save-summary", response_model=PatientTransaction)
async def save_summary_by_internal_id(internal_id: str, request: SaveSummaryRequest, user: dict = Depends(require_permission("session_summary"))) -> PatientTransaction:
    record = get_record_by_internal_id(internal_id)
    if record is None:
        raise HTTPException(status_code=404, detail="Patient not found")
    patient_id = record.patient_id
    today = datetime.utcnow().strftime("%Y-%m-%d")
    note = f"סיכום ביקור {today}"
    if request.doctor_name:
        note += f" | ד\"ר {request.doctor_name}"
    note += f"\n\n{request.summary}"
    visit_record = record.model_copy(update={"chief_complaint": note, "source": "visit"})
    tx = PatientTransaction(
        transaction_id=str(uuid.uuid4()),
        patient_id=patient_id,
        date=today,
        transaction_type="visit",
        raw_text=request.summary,
        extracted=visit_record,
    )
    save_transaction(tx)
    return tx
